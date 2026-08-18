import argparse
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.config import OUTPUT_DIR, POWERBI_WORKBOOK, ensure_directories
from src.db import get_engine


POWERBI_QUERIES = {
    "Fact_Loans": {
        "object": "credit_risk.mv_powerbi_loans",
        "grain": "One originated loan",
        "query": "SELECT * FROM credit_risk.mv_powerbi_loans ORDER BY loan_id",
    },
    "Fact_Thresholds": {
        "object": "credit_risk.vw_threshold_simulator",
        "grain": "One selected-model test threshold",
        "query": (
            "SELECT * FROM credit_risk.vw_threshold_simulator "
            "WHERE model_name = ("
            "SELECT model_name FROM credit_risk.loan_predictions LIMIT 1"
            ") ORDER BY threshold"
        ),
    },
    "Fact_ModelMetrics": {
        "object": "credit_risk.model_metrics",
        "grain": "One model, split, and threshold type",
        "query": (
            "SELECT * FROM credit_risk.model_metrics "
            "ORDER BY model_name, split, threshold_type"
        ),
    },
    "Fact_Portfolio": {
        "object": "credit_risk.portfolio_metrics",
        "grain": "One portfolio month",
        "query": "SELECT * FROM credit_risk.portfolio_metrics ORDER BY date",
    },
    "Fact_Stress": {
        "object": "credit_risk.vw_stress_scenarios",
        "grain": "One scenario and sector",
        "query": (
            "SELECT * FROM credit_risk.vw_stress_scenarios "
            "ORDER BY scenario, sector"
        ),
    },
    "Fact_Vintage": {
        "object": "credit_risk.vintage_analysis",
        "grain": "One vintage and month on books",
        "query": (
            "SELECT * FROM credit_risk.vintage_analysis "
            "ORDER BY vintage, months_on_books"
        ),
    },
    "Fact_Migration": {
        "object": "credit_risk.vw_rating_migration",
        "grain": "One year and sector",
        "query": (
            "SELECT * FROM credit_risk.vw_rating_migration "
            "ORDER BY year, sector"
        ),
    },
    "Feature_Importance": {
        "object": "credit_risk.feature_importance",
        "grain": "One selected-model feature",
        "query": (
            "SELECT * FROM credit_risk.feature_importance "
            "ORDER BY importance_mean DESC"
        ),
    },
}


def load_postgres_tables() -> dict[str, pd.DataFrame]:
    engine = get_engine()
    return {
        sheet_name: pd.read_sql(details["query"], engine)
        for sheet_name, details in POWERBI_QUERIES.items()
    }


def validate_tables(tables: dict[str, pd.DataFrame]) -> dict[str, int]:
    if set(tables) != set(POWERBI_QUERIES):
        raise ValueError(f"Unexpected Power BI tables: {sorted(tables)}")

    expected_exact = {
        "Fact_Loans": 50_000,
        "Fact_Portfolio": 120,
        "Fact_Stress": 60,
        "Fact_Vintage": 2_160,
    }
    counts = {name: len(frame) for name, frame in tables.items()}
    for name, expected in expected_exact.items():
        if counts[name] != expected:
            raise ValueError(f"{name}: expected {expected:,} rows, found {counts[name]:,}")
    if any(count <= 0 for count in counts.values()):
        raise ValueError(f"One or more Power BI tables are empty: {counts}")

    loans = tables["Fact_Loans"]
    required_loan_columns = {
        "loan_id",
        "defaulted",
        "data_split",
        "model_name",
        "default_probability",
        "operating_threshold",
        "risk_band",
        "underwriting_action",
        "term_risk_loss_proxy",
    }
    missing = required_loan_columns - set(loans.columns)
    if missing:
        raise ValueError(f"Fact_Loans is missing columns: {sorted(missing)}")
    if loans["loan_id"].duplicated().any():
        raise ValueError("Fact_Loans contains duplicate loan_id values.")
    if not loans["default_probability"].between(0, 1).all():
        raise ValueError("Fact_Loans contains probabilities outside [0, 1].")
    if loans[list(required_loan_columns)].isna().any().any():
        raise ValueError("Fact_Loans contains missing required analytical fields.")
    if set(loans["data_split"]) != {"train", "validation", "test"}:
        raise ValueError("Fact_Loans does not contain all temporal data splits.")

    thresholds = tables["Fact_Thresholds"]
    if set(thresholds["split"]) != {"test"}:
        raise ValueError("Fact_Thresholds must contain only the test split.")
    if thresholds["model_name"].nunique() != 1:
        raise ValueError("Fact_Thresholds must contain only the selected model.")
    if not thresholds["threshold"].between(0, 1).all():
        raise ValueError("Fact_Thresholds contains values outside [0, 1].")

    metrics = tables["Fact_ModelMetrics"]
    if not {"validation", "test"}.issubset(set(metrics["split"])):
        raise ValueError("Fact_ModelMetrics is missing validation or test metrics.")
    required_models = {"Logistic Regression", "Calibrated Random Forest", "XGBoost"}
    if not required_models.issubset(set(metrics["model_name"])):
        raise ValueError("Fact_ModelMetrics does not contain all three required models.")
    return counts


def build_manifest(counts: dict[str, int]) -> pd.DataFrame:
    rows = []
    for sheet_name, details in POWERBI_QUERIES.items():
        rows.append(
            {
                "sheet_name": sheet_name,
                "excel_table": f"tbl{sheet_name}",
                "postgresql_object": details["object"],
                "grain": details["grain"],
                "row_count": counts[sheet_name],
            }
        )
    return pd.DataFrame(rows)


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = False
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        max_row = worksheet.max_row
        max_column = worksheet.max_column
        if max_row >= 2 and max_column >= 1:
            table_name = f"tbl{worksheet.title}"
            table = Table(
                displayName=table_name,
                ref=f"A1:{worksheet.cell(max_row, max_column).coordinate}",
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)

        sample_end = min(max_row, 250)
        for column_cells in worksheet.iter_cols(
            min_row=1,
            max_row=sample_end,
            min_col=1,
            max_col=max_column,
        ):
            width = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(width + 2, 11), 36
            )

    workbook.save(path)


def write_workbook(tables: dict[str, pd.DataFrame], path: Path) -> dict[str, int]:
    counts = validate_tables(tables)
    manifest = build_manifest(counts)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        manifest.to_excel(writer, sheet_name="Manifest", index=False)
        for sheet_name, frame in tables.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
    format_workbook(path)
    return counts


def validate_workbook(path: Path, expected_counts: dict[str, int]) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    expected_sheets = ["Manifest", *POWERBI_QUERIES]
    if workbook.sheetnames != expected_sheets:
        raise ValueError(
            f"Workbook sheets do not match contract: {workbook.sheetnames}"
        )
    for sheet_name, expected_rows in expected_counts.items():
        observed_rows = workbook[sheet_name].max_row - 1
        if observed_rows != expected_rows:
            raise ValueError(
                f"{sheet_name}: workbook has {observed_rows:,} rows; "
                f"expected {expected_rows:,}."
            )
    workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Export PostgreSQL reporting objects into one centralized Excel "
            "workbook for Power BI."
        )
    )
    parser.parse_args()
    ensure_directories()

    # data/outputs is reserved for this one final Power BI source file.
    for generated_file in OUTPUT_DIR.iterdir():
        if generated_file.is_file() and generated_file.name != ".gitkeep":
            generated_file.unlink()

    temporary_path = POWERBI_WORKBOOK.with_name(
        f"{POWERBI_WORKBOOK.stem}.tmp{POWERBI_WORKBOOK.suffix}"
    )
    try:
        tables = load_postgres_tables()
        counts = write_workbook(tables, temporary_path)
        validate_workbook(temporary_path, counts)
        temporary_path.replace(POWERBI_WORKBOOK)
    except Exception:
        temporary_path.unlink(missing_ok=True)
        raise

    print(f"Power BI workbook saved to: {POWERBI_WORKBOOK}")
    print(f"Sheet row counts: {counts}")
    print("Use this one workbook as the only data source in Power BI Desktop.")


if __name__ == "__main__":
    main()
